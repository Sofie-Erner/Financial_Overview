# ----- Simplify Bank Statements -----
# This script will read through statements and specify their category
# Categories are based off csv file
# --------------------------------

# ----- Libraries -----
import openpyxl.utils
import openpyxl.utils.dataframe
import pandas as pd
import numpy as np
import datetime
import openpyxl
from openpyxl.styles import Alignment
import os
import sys
sys.path.append("../")

import src
from src.Additional_funcs import check_file
from src.Get_expense_categories import GetExpenseCategory

# ----- Function for simplifying bank statements -----
def SimplifyStatement(in_doc,exp_cat_doc,out_doc):
	path = os.path.abspath(os.getcwd()) # path to directory of script

	# ----- Variables -----
	out_date = "" # dates for bank statement
	out_cols = ["Date","Category","Money In","Money Out","Balance"] # columns of output file
	out_data = [] # list wil contain data for output file

	# ----- Dictionary with Categories for Expenses -----
	expenses = GetExpenseCategory(exp_cat_doc) # Dictionary with expense categories, will contain list of key words for each category
		
	exp_cat = expenses.keys() # List of expenses categories
	#print(expenses)
	
	# ----- Input File -----
	in_doc = str(path) + "/" + in_doc # path to bank statement (input file)
	empty_file = check_file(in_doc,"xlsx")

	if empty_file == 1:
		print("Error: input statement,",in_doc,", is empty")
		exit()

	f_in = pd.ExcelFile(in_doc, engine='openpyxl') 
	#print(f_in.sheet_names)

	# ----- Common Expressions to Replace
	common_expr = np.array(["card payment to ","credit from ","direct debit payment to ","bill payment via faster payment to "])

	# ----- Import Data ----- 
	df = pd.read_excel(in_doc)

	df.dropna(how="all",axis=1,inplace=True) #remove empty columns

	if len(df.columns) != 5:
		print("Error in number of columns")
		exit()

	# --- rename columns
	df.rename({df.columns[0]:"Date"},axis='columns', inplace=True)
	df.rename({df.columns[1]:"Description"},axis='columns', inplace=True)
	df.rename({df.columns[2]:"Money in"},axis='columns', inplace=True)
	df.rename({df.columns[3]:"Money out"},axis='columns', inplace=True)
	df.rename({df.columns[4]:"Balance"},axis='columns', inplace=True)

	# --- make descriptions easier to deal with
	df["Description"] = df["Description"].str.lower()
	for expr in common_expr:
		df.replace(expr,"", inplace=True, regex=True)
		
	# ----- Loop Over Data ----- 
	for i in range(len(df)):
		# --- Get Month
		if "XXXX" in str(df.loc[i,"Date"]):
			dates = df.loc[i,"Description"].split(" ")
			date1 = dates[0].split("/")
			date2 = dates[2].split("/")

			out_date = '-'.join(date1) + "_" + "-".join(date2)

		elif df.loc[i,"Date"] == "Date": # don't do anything for this line
			continue

		elif not pd.isnull(df.loc[i,"Balance"]): # For lines with expenses
			data1 = df.loc[i].values.flatten().tolist() # get dataframe row
			data1[0] = str(data1[0].day) + " " + str(data1[0].month) + " " + str(data1[0].year) #date

			# Find expense category
			alloc = 0 # whether it's been allocated

			for cat in expenses: # for each category
				if any(word in data1[1] for word in expenses[cat]): # if description in categories
					data1[1] = cat # note category
					alloc = 1

			if alloc == 0: # if expense category was not found
				if not pd.isnull(df.loc[i,"Money in"]):
					data1[1] = "income"
				else:
					data1[1] = "unknown"

			out_data.append(data1) # add to data

	# ----- Output File -----
	out_doc = str(path) + "/" + out_doc # path to output simplified bank statement
	empty_file = check_file(out_doc,"xlsx") # check file

	if empty_file == 0: # if file exists
		print("Output file ",out_doc," is not empty")

		txt_in = input("The file will be appended, want to continue? [y/n]")
		if txt_in != "y":
			print("'No' selected, exiting")
			exit()

		out_workbook = openpyxl.load_workbook(filename=out_doc) # open file with append
	else: # if file does not exist
		out_workbook = openpyxl.Workbook() # open new excel file


	out_sheets = out_workbook.sheetnames # get sheets

	# Check if dates already in file
	if out_date in out_sheets:
		print("Dates ",out_date," already in file.")
		txt_in = input("The content will be overwritten, want to continue? [y/n]")
		if txt_in != "y":
			print("'No' selected, exiting")
			exit()

	del out_workbook[out_date] # remove sheet to be re-written
	out_worksheet = out_workbook.create_sheet(out_date) # create new sheet
	out_sheets = out_workbook.sheetnames # update sheets
	
	# Add new sheet 
	out_workbook.active = out_workbook.sheetnames.index(out_date) # change active sheet to newly added
	out_worksheet = out_workbook.active
	
	# ---- Output dataframe
	df_out = pd.DataFrame(data=out_data,columns=out_cols) # make data into dataframe

	# Change date format
	dates = df_out["Date"].tolist()
	dates = [ "/".join(dates[i].split(" ")) for i in range(0,len(dates)) ]
	df_out.update(pd.DataFrame({'Date':dates}))
	
	for r in openpyxl.utils.dataframe.dataframe_to_rows(df_out, index=True, header=True): #add dataframe
		out_worksheet.append(r)
	
	# remove extra rows and columns inserted
	out_worksheet.delete_rows(2) # delete empty second row
	out_worksheet.delete_cols(1) # delete first column with row numbers

	# ----- Merge balance for dates which are the same
	date_counts = [(i,dates.count(dates[i])) for i in range(0,len(dates))]

	col_ix = df_out.columns.get_loc("Balance")  + 1 # index from 1 

	# merge cells with the same date (dates assumed to be ordered )
	i = 0
	while i < len(date_counts):
		id = date_counts[i][0] + 2 # indexing from 1 plus column heads
		count = date_counts[i][1]

		if count > 1: # more than one 
			# merge "balance" cell for same dates, have value for last expense
			out_worksheet.merge_cells(start_row=id,start_column=col_ix,end_row=id+count-1,end_column=col_ix)
			cell = out_worksheet.cell(row=id,column=col_ix) # get merged cell
			cell.value = str(df_out.loc[id+count-3,"Balance"]) # remove extra rows, assign correct balance
			cell.alignment = Alignment(vertical='center',horizontal='right')

			i = i + count # move to next cell after merge
		else:
			i += 1 # no merge


	print("Bank statement for ",out_date," added")		
	out_workbook.save(out_doc) # save and close file